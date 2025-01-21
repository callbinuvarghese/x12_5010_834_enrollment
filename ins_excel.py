import os
from loguru import logger
from typing import List

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet 
from openpyxl.comments import Comment
from openpyxl.cell import Cell
from openpyxl.styles import Color, PatternFill, Font

from ins_class import INS

def create_excel_spreadsheet_openpyxl(data: List[INS], filename: str = "output.xlsx"):
    """Creates an Excel spreadsheet from a list of dictionaries using openpyxl.

    Args:
        data: A list of dictionaries, where each dictionary represents a row.
        filename: The name of the Excel file to create (default: "output.xlsx").
    """
    if not data:
        print("No data provided to create the spreadsheet.")
        return

    try:
        wb = Workbook()
        ws = wb.active

        # Add header row (keys from the first dictionary)
        header = list(data[0].keys()) if data else []
        ws.append(header)

        # Add data rows
        for row_data in data:
            row = list(row_data.values())
            ws.append(row)

        wb.save(filename)
        print(f"Excel spreadsheet created successfully: {filename}")
    except Exception as e:
        print(f"An error occurred while creating the spreadsheet: {e}")

def create_excel_spreadsheet_openpyxl_df(df, filename: str = "output.xlsx"):
    """Creates an Excel spreadsheet from a pandas dataframe using openpyxl.

    Args:
        df: A pandas dataframe
        filename: The name of the Excel file to create (default: "output.xlsx").
    """
    if df.empty:
        print("No data provided to create the spreadsheet.")
        return

    try:
        wb = Workbook()
        ws = wb.active

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        wb.save(filename)
        print(f"Excel spreadsheet created successfully: {filename}")
    except Exception as e:
        print(f"An error occurred while creating the spreadsheet: {e}")


def create_workbook(filename:str)->Workbook:
    # Create a new workbook
    wb:Workbook = None
    if (os.path.isfile(filename) and  os.access(filename, os.W_OK)):
        logger.info(f"Existing workbook filename:{filename}.. will be overwritten")
    else:
        logger.info(f"Creating an new workbook for filename:{filename}")
    wb = Workbook() 
    return wb

def create_worksheet(wb:Workbook, sheetname:str)->Worksheet:
    # Create a new worksheet
    sheet:Worksheet= None
    if sheetname not in wb.sheetnames:
        logger.info(f"Creating an new sheet for sheetname:{sheetname} in workbook")
        sheet=wb.create_sheet(sheetname)
    else:
        logger.info(f"Selecting an existing sheet for sheetname:{sheetname} in workbook")
        sheet= wb[sheetname]
    return sheet

def setcellhead(cell:Cell):
    cell.fill=PatternFill(patternType='solid',
                                        fill_type='solid', 
                                        fgColor=Color('C4C4C4'))
    ft = Font(color="000000FF",bold=True)
    cell.font=ft

def add_sheetdata_INS(ws:Worksheet, list:List[INS], rowstart:int)->int:
    rowstart+=1
    col=0
    cell=ws.cell(row=rowstart, column=col+1, value="RelID")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+2, value="Yes/No")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+3, value="Reltn")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+4, value="MaintType")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+5, value="MaintReason")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+6, value="BenStatus")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+7, value="Medicare")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+8, value="OccLength")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+9, value="Handicap")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+10, value="")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+11, value="IdCode")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+12, value="IdCodeQual")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+13, value="LastNameOrOrg")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+14, value="FirstName")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+15, value="MiddleName")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+16, value="NMCdQual")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+17, value="NMCd")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+18, value="EntRelCd")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+19, value="EntIdCd")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+20, value="ContFnCd")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+21, value="ContEnuCd")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+22, value="ContCommQual")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+23, value="ContCommDetail")
    setcellhead(cell)    
    cell=ws.cell(row=rowstart, column=col+25, value="Addr1")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+26, value="Addr2")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+27, value="City")
    setcellhead(cell)   
    cell=ws.cell(row=rowstart, column=col+28, value="State")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+29, value="Zip")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+30, value="Country")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+31, value="LocId")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+33, value="DOB")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+34, value="M/F")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+35, value="Race")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+37, value="MtRsnCd")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+38, value="SrcCd")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+39, value="CvrgCd")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+40, value="EmpStCd")
    setcellhead(cell)


    for ins in list:
        rowstart+=1
        col=0
        cell=ws.cell(row=rowstart, column=col+1, value=rowstart)
        cell=ws.cell(row=rowstart, column=col+2, value=ins.yes_no_response_code)
        cell=ws.cell(row=rowstart, column=col+3, value=ins.dependent_code)
        cell=ws.cell(row=rowstart, column=col+4, value=ins.maintenance_type_code)
        cell=ws.cell(row=rowstart, column=col+5, value=ins.maintenance_reason_code)
        cell=ws.cell(row=rowstart, column=col+6, value=ins.benefit_status_code)
        cell=ws.cell(row=rowstart, column=col+7, value=ins.medicare_status_code)
        cell=ws.cell(row=rowstart, column=col+8, value=ins.occ_length_code)
        cell=ws.cell(row=rowstart, column=col+9, value=ins.handicap_ind)
        cell=ws.cell(row=rowstart, column=col+10, value="")
        cell=ws.cell(row=rowstart, column=col+11, value=ins.nm1_segment.entity_identifier_code)
        cell=ws.cell(row=rowstart, column=col+12, value=ins.nm1_segment.entity_type_qualifier)
        cell=ws.cell(row=rowstart, column=col+13, value=ins.nm1_segment.name_last_or_organization_name)
        cell=ws.cell(row=rowstart, column=col+14, value=ins.nm1_segment.name_first)
        cell=ws.cell(row=rowstart, column=col+15, value=ins.nm1_segment.name_middle)
        cell=ws.cell(row=rowstart, column=col+16, value=ins.nm1_segment.identification_code_qualifier)
        cell=ws.cell(row=rowstart, column=col+17, value=ins.nm1_segment.identification_code)
        cell=ws.cell(row=rowstart, column=col+18, value=ins.nm1_segment.entity_relationship_code)
        cell=ws.cell(row=rowstart, column=col+19, value=ins.nm1_segment.entity_identifier_code_2)
        if ins.per_segment:
            cell=ws.cell(row=rowstart, column=col+20, value=ins.per_segment.contact_function_code)
            cell=ws.cell(row=rowstart, column=col+21, value=ins.per_segment.contact_enumeration_code)
            cell=ws.cell(row=rowstart, column=col+22, value=ins.per_segment.contact_communication_number_qualifier)
            cell=ws.cell(row=rowstart, column=col+23, value=ins.per_segment.contact_communication_number)
        cell=ws.cell(row=rowstart, column=col+25, value=ins.n3_segment.address_information_1)
        cell=ws.cell(row=rowstart, column=col+26, value=ins.n3_segment.address_information_2)
        cell=ws.cell(row=rowstart, column=col+27, value=ins.n4_segment.city_name)
        cell=ws.cell(row=rowstart, column=col+28, value=ins.n4_segment.state_or_province_code)
        cell=ws.cell(row=rowstart, column=col+29, value=ins.n4_segment.postal_code)
        cell=ws.cell(row=rowstart, column=col+30, value=ins.n4_segment.country_code)
        cell=ws.cell(row=rowstart, column=col+31, value=ins.n4_segment.location_identifier)
        cell=ws.cell(row=rowstart, column=col+33, value=ins.dmg_segment.date_time_period)
        cell=ws.cell(row=rowstart, column=col+34, value=ins.dmg_segment.gender_code)
        cell=ws.cell(row=rowstart, column=col+35, value=ins.dmg_segment.race_or_ethnicity_code)
        cell=ws.cell(row=rowstart, column=col+37, value=ins.hd_segment.maintenance_reason_code)
        cell=ws.cell(row=rowstart, column=col+38, value=ins.hd_segment.maintenance_type_code)
        cell=ws.cell(row=rowstart, column=col+39, value=ins.hd_segment.plan_coverage_description)
        cell=ws.cell(row=rowstart, column=col+40, value=ins.hd_segment.employee_status_code)

def add_sheetdata_INS_REF(ws:Worksheet, list:List[INS], rowstart:int, refrow:int)->int:
    rowstart+=1
    col=0
    cell=ws.cell(row=rowstart, column=col+1, value="RefRow")
    setcellhead(cell)

    cell=ws.cell(row=rowstart, column=col+2, value="LastOrOrg")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+3, value="FirstName")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+4, value="Mid")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+5, value="DOB")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+6, value="ID")
    setcellhead(cell)


    cell=ws.cell(row=rowstart, column=col+8, value="IDQual")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+9, value="ID")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+10, value="Desc")
    setcellhead(cell)

    for ins in list:
        refrow+=1
        for ref in ins.ref_segments:
            col=0
            rowstart+=1
            cell=ws.cell(row=rowstart, column=col+1, value=refrow)

            cell=ws.cell(row=rowstart, column=col+2, value=ins.nm1_segment.name_last_or_organization_name)
            cell=ws.cell(row=rowstart, column=col+3, value=ins.nm1_segment.name_first)
            cell=ws.cell(row=rowstart, column=col+4, value=ins.nm1_segment.name_middle)
            cell=ws.cell(row=rowstart, column=col+5, value=ins.dmg_segment.date_time_period)
            cell=ws.cell(row=rowstart, column=col+6, value=ins.nm1_segment.identification_code)

            cell=ws.cell(row=rowstart, column=col+8, value=ref.reference_identification_qualifier)
            cell=ws.cell(row=rowstart, column=col+9, value=ref.reference_identification)
            cell=ws.cell(row=rowstart, column=col+10, value=ref.description)

  
def add_sheetdata_INS_DTP(ws:Worksheet, list:List[INS], rowstart:int, refrow:int)->int:
    rowstart+=1
    col=0
    cell=ws.cell(row=rowstart, column=col+1, value="RefRow")
    setcellhead(cell)

    cell=ws.cell(row=rowstart, column=col+2, value="LastOrOrg")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+3, value="FirstName")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+4, value="Mid")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+5, value="DOB")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+6, value="ID")
    setcellhead(cell)


    cell=ws.cell(row=rowstart, column=col+8, value="IDQual")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+9, value="ID")
    setcellhead(cell)
    cell=ws.cell(row=rowstart, column=col+10, value="Desc")
    setcellhead(cell)

    for ins in list:
        refrow+=1
        for ref in ins.dtp_segments:
            col=0
            rowstart+=1
            cell=ws.cell(row=rowstart, column=col+1, value=refrow)

            cell=ws.cell(row=rowstart, column=col+2, value=ins.nm1_segment.name_last_or_organization_name)
            cell=ws.cell(row=rowstart, column=col+3, value=ins.nm1_segment.name_first)
            cell=ws.cell(row=rowstart, column=col+4, value=ins.nm1_segment.name_middle)
            cell=ws.cell(row=rowstart, column=col+5, value=ins.dmg_segment.date_time_period)
            cell=ws.cell(row=rowstart, column=col+6, value=ins.nm1_segment.identification_code)

            cell=ws.cell(row=rowstart, column=col+8, value=ref.date_time_qualifier)
            cell=ws.cell(row=rowstart, column=col+9, value=ref.date_time_format_qualifier)
            cell=ws.cell(row=rowstart, column=col+10, value=ref.date_time_period)


def create_excel(list:List[INS]):
    filename="ins.xlsx"
    wb=create_workbook(filename)
    # Delete the default sheet
    std = wb["Sheet"]
    wb.remove(std)

    ws1=create_worksheet(wb, "INS")
    add_sheetdata_INS(ws1, list, 0)

    ws2=create_worksheet(wb, "INS-REF")
    add_sheetdata_INS_REF(ws2, list, 0,0)

    ws3=create_worksheet(wb, "INS-DTP")
    add_sheetdata_INS_DTP(ws3, list, 0, 0)

    wb.save(filename)
    logger.info(f"Excel spreadsheet created successfully: {filename}")
